import tkinter as tk
from tkinter import messagebox
from datetime import datetime
from datetime import date 
from tkcalendar import Calendar  # type: ignore
import shutil
import os
from tkinter import ttk
from openpyxl.styles import PatternFill
from datetime import date
import pyvisa 
import keyboard
import openpyxl
import time
import string

# Initialize the resource manager
rm = pyvisa.ResourceManager()

# Open the resource
multimeter = rm.open_resource('GPIB0::8::INSTR')
powerS = rm.open_resource('GPIB0::1::INSTR')
loadS = rm.open_resource('GPIB0::2::INSTR')

# Initialize the instrument
multimeter.write("*RST")
powerS.write("*RST")
loadS.write("*RST")

# For the sheet
wb = openpyxl.load_workbook(r"BatteryCharger.xlsx")
sheet = wb.active 

c1 = sheet.cell(row = 18, column = 5)

def set_input_voltage(voltage):
	powerS.write("ACDC AC") #Set output to AC
	powerS.write("FSET 60") #Set output to 60Hz
	powerS.write("RANGE 200") #Set range to 200V mode
	powerS.write("ACVSET {}".format(voltage)) #Set output voltage to "voltage"
	powerS.write("ACILIM 5.0") #Set output current limit to 5A

def get_output_voltage():
	multimeter.write(":SENSe:FUNCtion 'VOLTage:DC'")  # Set the function to DC voltage
	multimeter.write(":SENSe:VOLTage:DC:RANGe:AUTO ON")  # Adjusts the range automatically
	vForMulti = multimeter.query(":READ?")
	vForMulti = vForMulti[:13] + vForMulti[14:]
	return float(vForMulti)

def get_input_voltage():
	powerSV = powerS.query("VOUT?")
	powerSV = powerSV[:13] + powerSV[14:]
	return float(powerSV)

def get_input_current():
	powerSI = powerS.query("IOUT?")
	powerSI = powerSI[:13] + powerSI[14:]
	return float(powerSI)

def col2num(col):
	num = 0
	for c in col:
			if c in string.ascii_letters:
					num = num * 26 + (ord(c.upper()) - ord('A')) + 1
	return num

def write_value_to_cell(value, cell):
	cell_col = col2num(''.join(filter(lambda x : x in string.ascii_letters, cell)))
	cell_row = int(''.join(filter(lambda x : x in string.digits, cell)))
	c1 = sheet.cell(row = cell_row, column = cell_col)
	c1.value = value
	wb.save(r"BatteryCharger.xlsx")

class CommandGUI:
	def __init__(self, master):
		self.master = master
		self.master.title("Command Entry")
		self.master.geometry("1200x400")
		self.language = None
		self.command_index = 0
		self.commands = {
			'English': [
				"Enter The Orderer of Product: Hit Enter to record",
				"Hit Enter to Choose a Date",
				"Enter Drawing Number: Hit Enter to record",
				"Enter Room Temperature: Hit Enter to record",
				"Enter Charger Serial Number: Hit Enter to record",
				"Enter Official Format: Hit Enter to record",
				"Enter Serial Number: Hit Enter to record",
				"Enter Confirmer Name: Hit Enter to record",
				"Enter Tester Name: Hit Enter to record",
				"Connect R and S terminals to 200V power supply\nConnect Electronic Load to N11 and P11\nConnect multimeter probes to the two terminals above the MCCBR breaker\nPress Enter to continue",
				"Turn on MCCBR breaker. PL light on back of the power supply should be on\nPress Enter if PL light is on",
				"Turn on MCCB0 and MCCB11 breakers\nPress Enter to continue",
				"Use screwdriver to adjust the \n F potentiometer on the charger to reach approximately 26.8V on the load\nPress Enter to continue",
				"Adjust F potentiometer using a screwdriver to reach 26.8V on the load",
				"Again, Adjust F potentiometer using a screwdriver to reach 26.8V on the load",
				"For the last time, adjust F potentiometer using a screwdriver to reach 26.8V on the load",
				"Turn CA Potentiometer to maximum (to the right)",
				"Hit Enter, wait for the Load to be in CR mode (about 15 sec)\nthen turn potentiometer until the multimeter shows 24V",
				"Turn F and VA Potentiometers to the right to measure maximum voltage\nPress Enter to continue",
				"Turn F and VA Potentiometers to the left to measure minimum voltage\nPress Enter to continue",
				"Hit Enter"
			],
			'Japanese': [
				"商品の注文者を入力してください： 記録するにはEnterキーを押してください",
				"日付を選択するにはEnterキーを押してください",
				"図面番号を入力してください",
				"室温を入力してください：\nEnter キーを押して記録します",
				"チャージャーのシリアルナンバーを入力してください",
				"公式フォーマットを入力してください：\nEnter キーを押して記録",
				"シリアルナンバーを入力してください",
				"確認者名を入力してください：\nEnter を押して記録",
				"テスター名を入力してください：\nEnter を押して記録します",
				"R端子とS端子を200V電源に接続する\n電子負荷をN11とP11に接続する\nマルチメーターのプローブをMCCBRブレーカーの上の2つの端子に接続します\nEnterを押して続ける",
				"MCCBRブレーカをオンにする\n電源装置の背面にあるPLランプが点灯しているはずです\nPL ランプが点灯している場合は、Enter を押す",
				"MCCB0 および MCCB11 ブレーカをオンにする\nEnterを押して続行します",
				"負荷が約 26.8V になるように、ドライバーを使用して充電器の F ポテンショメータを調整します\nEnterを押して続行",
				"負荷で26.8VになるようにドライバーでFポテンショメーターを調整する",
				"もう一度、負荷が26.8VになるようにドライバーでFポテンショメーターを調整してください",
				"最後に、ドライバーでFポテンショメーターを調整し、\n負荷が26.8Vになるようにします",
				"CAポテンショメーターを最大（右側）に回します",
				"ロードがCRモードになるのを待ち（約15秒）、\nマルチメーターが24Vを示すまでポテンショメーターを回します",
				"FポテンショメーターとVAポテンショメーターを右に回して最大電圧を測定します\nEnterキーを押して続けます",
				"FポテンショメーターとVAポテンショメーターを左に回して最小電圧を測定します\nEnterキーを押して続けます",
				"エンターキーを押す"
			]
		}
		# rm = pyvisa.ResourceManager()

		self.data = {}
		self.show_language_selection()

	def show_language_selection(self):
		self.label = tk.Label(self.master, text="Select Language / 言語を選択してください", font=("Arial", 20))
		self.label.pack(pady=20)

		self.english_button = tk.Button(self.master, text="English", font=("Arial", 20), command=lambda: self.set_language('English'))
		self.english_button.pack(pady=10)

		self.japanese_button = tk.Button(self.master, text="日本語", font=("Arial", 20), command=lambda: self.set_language('Japanese'))
		self.japanese_button.pack(pady=10)

	def set_language(self, language):
		self.language = language
		self.label.pack_forget()
		self.english_button.pack_forget()
		self.japanese_button.pack_forget()
		self.setup_gui()

	def setup_gui(self):
		self.label = tk.Label(self.master, text=self.commands[self.language][self.command_index], font=("Arial", 20))
		self.label.pack(pady=20)

		self.entry = tk.Entry(self.master, font=("Arial", 20))
		self.entry.pack(pady=10)
		self.entry.bind('<Return>', self.on_enter)

	def on_enter(self, event):
		current_command = self.commands[self.language][self.command_index]
		input_value = self.entry.get()
		wb.save(r"BatteryCharger.xlsx")
		self.entry.delete(0, tk.END)
		self.command_index += 1
		# Process input based on the current command and ADD TO SS PER RESPONSE
		if "Orderer" in current_command or "注文者" in current_command:
			self.data['orderer'] = input_value
			c1 = sheet.cell(row=3, column=5)  
			c1.value = input_value 

		elif "Date" in current_command or "日付" in current_command:
			today_bad = date.today()
			today = today_bad.strftime('%Y-%m-%d')
			year = today[:4]
			month = today[5:7]
			day = today[8:]
			write_value_to_cell(year,'E6')
			write_value_to_cell(month,'H6')
			write_value_to_cell(day,'K6')
		elif "Drawing Number" in current_command or "図面番号" in current_command:
			self.data['drawing_number'] = input_value
			c1 = sheet.cell(row=7, column=5)  
			c1.value = input_value 
		elif "Temperature" in current_command or "室温" in current_command:
			self.data['room_temperature'] = input_value
			c1 = sheet.cell(row=8, column=5)  
			c1.value = input_value 
		elif "Charger Serial Number" in current_command or "充電器シリアル番号" in current_command:
			self.data['charger_serial_number'] = input_value
			c1 = sheet.cell(row=9, column=5)  
			c1.value = input_value 
		elif "Official Format" in current_command or "正式形式" in current_command:
			self.data['official_format'] = input_value
			c1 = sheet.cell(row=5, column=19)  
			c1.value = input_value 
		elif "Serial Number" in current_command or "シリアル番号" in current_command:
			self.data['serial_number'] = input_value
			c1 = sheet.cell(row=6, column=19)  
			c1.value = input_value 
		elif "Confirmer Name" in current_command or "確認者名" in current_command:
			self.data['confirmer_name'] = input_value
			c1 = sheet.cell(row=7, column=19)  
			c1.value = input_value 
		elif "Tester Name" in current_command or "テスター名" in current_command:
			self.data['tester_name'] = input_value
			c1 = sheet.cell(row=8, column=19)  
			c1.value = input_value 



		if self.command_index < len(self.commands[self.language]):
			self.label.config(text=self.commands[self.language][self.command_index])
			# Run internal code between screens
			self.run_internal_code()

		else:
			self.finish()



	def show_calendar(self):
		self.calendar_window = tk.Toplevel(self.master)
		self.calendar_window.title("Select Date")
		self.calendar = Calendar(self.calendar_window, selectmode="day", date_pattern="yyyy-mm-dd")
		self.calendar.pack(pady=20)

		select_button = tk.Button(self.calendar_window, text="Select", command=self.select_date)
		select_button.pack(pady=10)

	def select_date(self):
		selected_date = self.calendar.get_date()
		self.data['date'] = selected_date


		year = selected_date[0:4]
		c1 = sheet.cell(row=6, column=5)  
		c1.value = year 

		month = selected_date[5:7]

		c1 = sheet.cell(row=6, column=8)  
		c1.value = month 

		day = selected_date[8:]
		c1 = sheet.cell(row=6, column=11)  
		c1.value = day


		self.calendar_window.destroy()
		self.entry.delete(0, tk.END)
		self.command_index += 1
		if self.command_index < len(self.commands[self.language]):
			self.label.config(text=self.commands[self.language][self.command_index])
			# Run internal code between screens
			self.run_internal_code()
		else:
			self.finish()

	def run_internal_code(self):
		self.show_loading_screen()

	def show_loading_screen(self):
		if self.command_index >= 10:
			self.entry.pack_forget()
			self.loading_window = tk.Toplevel(self.master)
			self.loading_window.title("Processing")
			self.loading_window.geometry("800x200")

			if(self.language == 'English'):
				label = tk.Label(self.loading_window, text="Processing...", font=("Arial", 20))
				label.pack(pady=20)
			else:
				label = tk.Label(self.loading_window, text="続行する前に、この画面が消えるのを待つ", font=("Arial", 20))
				label.pack(pady=20)

			progress = ttk.Progressbar(self.loading_window, orient='horizontal', 	 
			mode='indeterminate', length=300)
			progress.pack(pady=20)
			progress.start()

			self.loading_window.after(1000, self.execute_internal_code)  # Simulate code execution delay

	def execute_internal_code(self):
		##### Internal Code Start #####
		if self.command_index == 10:
			print("Running internal code for step 10")
			
			set_input_voltage(200)
			powerS.write("OUT ON") #Set output ON
			time.sleep(5)
			outPowerSCurr = powerS.query("IOUT?") #Query about output current
			vForMulti = get_output_voltage()

			#c1 = sheet.cell(row = 12, column = 9)
			#c1.value = vForMulti

			#FOR OUTPUT VOLTAGE
			# vPower = float(powerS.query(":VOLTage?"))
			# while vPower < 200.0:
			# 	vPower = float(powerS.query(":VOLTage?"))

		elif self.command_index == 11:			
			print("Running internal code for step 11")
		elif self.command_index == 12:
			print("Running internal code for step 12")
		elif self.command_index == 13:
			print("Running internal code for step 13")
			
			# READ 1

			# Turn On Load
			loadS.write("CURRent 10")
			loadS.write("INPut ON")
			loadS.write("OUTput ON")
			time.sleep(5)

			# Record Input Voltage 
			powerSV = get_input_voltage()
			write_value_to_cell(powerSV, "E18")

			# Record Input Current
			powerSI = get_input_current()
			write_value_to_cell(powerSI, "G18")

			# Record Input True Power
			sourcePowerWatt = powerS.query("WATT?")
			sourcePowerWatt = sourcePowerWatt[:13] + sourcePowerWatt[14:]
			sourcePowerWatt = float(sourcePowerWatt)
			c1 = sheet.cell(row = 18, column = 9)
			c1.value = sourcePowerWatt

			# READ 2

			# Record Output Voltage
			multimeter.write(":SENSe:FUNCtion 'VOLTage:DC'")  # Set the function to DC voltage
			multimeter.write(":SENSe:VOLTage:DC:RANGe:AUTO ON")  # Set the range to 10V, adjust as necessary
			time.sleep(2)
			vForMulti = multimeter.query(':READ?')
			vForMulti = vForMulti[:13] + vForMulti[14:]
			vForMulti = float(vForMulti)
			#print("vForMulti is " + vForMulti + "V")
			c1 = sheet.cell(row = 18, column = 11)
			c1.value = vForMulti

			# Record Current
			aForMulti = loadS.query("MEASure:CURRent?")
			time.sleep(2)
			aForMulti = aForMulti[:13] + aForMulti[14:]
			aForMulti = float(aForMulti)
			#print("aForMulti is " + aForMulti + "A")
			c1 = sheet.cell(row = 18, column = 13)
			c1.value = aForMulti

			# Calculate and Record True Power
			powerRead2 = vForMulti * aForMulti #THIS IS POWER FOR READ 2
			#print("powerRead2 is " + powerRead2 + "W")
			c1 = sheet.cell(row = 18, column = 15)
			c1.value = powerRead2

			#READ 3

			# Calculate and Record Power Efficiency
			powerEfficiency = powerRead2 / sourcePowerWatt * 100 
			#print("powerEfficiency is " + powerEfficiency + "%")
			c1 = sheet.cell(row = 18, column = 17)
			c1.value = powerEfficiency #THIS IS FOR READ 3

			#READ 4

			# Calculate and Record Power Factor
			sourcePowerFactor = powerS.query("PF?") #POWER FACTOR
			sourcePowerFactor = float(sourcePowerFactor)
			sourcePowerFactor = sourcePowerFactor * 100
			c1 = sheet.cell(row = 18, column = 20) #THIS IS FOR READ 4
			c1.value = sourcePowerFactor

			# Turn Off Load
			loadS.write("INPut OFF")
			loadS.write("OUTPut OFF")

			# Input = 200V
			powerS.write("ACDC AC") #Set output to AC
			powerS.write("FSET 60") #Set output to 50Hz
			powerS.write("RANGE 200") #Set range to 200V mode
			powerS.write("ACVSET 200") #Set output voltage to 200V
			powerS.write("ACILIM 5.0") #Set output current limit to 8A
			powerS.write("OUT ON") #Set output ON
			
			outPowerSCurr = powerS.write("IOUT?") #Query about output current

		elif self.command_index == 14:

			print("Running internal code for step 14")
			
			time.sleep(2)

			# Record Output Voltage
			multimeter.write(":SENSe:FUNCtion 'VOLTage:DC'")  # Set the function to DC voltage
			multimeter.write(":SENSe:VOLTage:DC:RANGe:AUTO ON")  # Set the range to 10V, adjust as necessary
			x1y2 = multimeter.query(":READ?")
			x1y2 = x1y2[:13] + x1y2[14:]
			x1y2 = float(x1y2)
			#print(x1y2 + "V")
			c1 = sheet.cell(row = 27, column = 5)
			c1.value = x1y2

			# Input = 180V
			powerS.write("ACDC AC") #Set output to DC
			powerS.write("RANGE 200") #Set range to 200V mode
			powerS.write("ACVSET 180") #Set output voltage to 200V
			powerS.write("ACILIM 5.0") #Set output current limit to 8A
			powerS.write("OUT ON") #Set output ON
			outPowerSCurr = powerS.write("IOUT?") #Query about output current

		elif self.command_index == 15:

			print("Running internal code for step 15")
			
			time.sleep(2)

			# Record Output Voltage
			multimeter.write(":SENSe:FUNCtion 'VOLTage:DC'")  # Set the function to DC voltage
			multimeter.write(":SENSe:VOLTage:DC:RANGe:AUTO ON")  # Set the range to 10V, adjust as necessary
			x1y2 = multimeter.query(":READ?")
			x1y2 = x1y2[:13] + x1y2[14:]
			x1y2 = float(x1y2)
			#print(x1y2 + "V")
			c1 = sheet.cell(row = 26, column = 5)
			c1.value = x1y2

			# Input = 220V
			powerS.write("ACDC AC") #Set output to DC
			powerS.write("RANGE 200") #Set range to 200V mode
			powerS.write("ACVSET 220") #Set output voltage to 200V
			powerS.write("ACILIM 5.0") #Set output current limit to 8A
			powerS.write("OUT ON") #Set output ON
			outPowerSCurr = powerS.write("IOUT?") #Query about output current

		elif self.command_index == 16:

			print("Running internal code for step 16")
			
			time.sleep(2)

			# Record Output Voltage
			multimeter.write(":SENSe:FUNCtion 'VOLTage:DC'")  # Set the function to DC voltage
			multimeter.write(":SENSe:VOLTage:DC:RANGe:AUTO ON")  # Set the range to 10V, adjust as necessary
			x1y2 = multimeter.query(":READ?")
			x1y2 = x1y2[:13] + x1y2[14:]
			x1y2 = float(x1y2)
			#print(x1y2 + "V")
			c1 = sheet.cell(row = 28, column = 5)
			c1.value = x1y2

			# Input = 200V
			powerS.write("ACDC AC") #Set output to DC
			powerS.write("RANGE 200") #Set range to 200V mode
			powerS.write("ACVSET 200") #Set output voltage to 200V
			powerS.write("ACILIM 5.0") #Set output current limit to 8A
			powerS.write("OUT ON") #Set output ON
			outPowerSCurr = powerS.write("IOUT?") #Query about output current

			loadS.write("INPut ON")
			loadS.write("OUTput ON")

			loadS.write("CURRent 5")

			time.sleep(5)

			# Record Output Voltage
			multimeter.write(":SENSe:FUNCtion 'VOLTage:DC'")  # Set the function to DC voltage
			multimeter.write(":SENSe:VOLTage:DC:RANGe:AUTO ON")  # Set the range to 10V, adjust as necessary
			x1y2 = multimeter.query(":READ?")
			x1y2 = x1y2[:13] + x1y2[14:]
			x1y2 = float(x1y2)
			#print(x1y2 + "V")
			c1 = sheet.cell(row = 27, column = 8)
			c1.value = x1y2

			# Input = 180V
			powerS.write("ACDC AC") #Set output to DC
			powerS.write("RANGE 200") #Set range to 200V mode
			powerS.write("ACVSET 180") #Set output voltage to 200V
			powerS.write("ACILIM 5.1") #Set output current limit to 8A
			powerS.write("OUT ON") #Set output ON
			outPowerSCurr = powerS.write("IOUT?") #Query about output current

			time.sleep(5)

			# Record Output Voltage
			multimeter.write(":SENSe:FUNCtion 'VOLTage:DC'")  # Set the function to DC voltage
			multimeter.write(":SENSe:VOLTage:DC:RANGe:AUTO ON")  # Set the range to 10V, adjust as necessary
			x1y2 = multimeter.query(":READ?")
			x1y2 = x1y2[:13] + x1y2[14:]
			x1y2 = float(x1y2)
			#print(x1y2 + "V")
			c1 = sheet.cell(row = 26, column = 8)
			c1.value = x1y2

			# Input = 220V
			powerS.write("ACDC AC") #Set output to DC
			powerS.write("RANGE 200") #Set range to 200V mode
			powerS.write("ACVSET 220") #Set output voltage to 200V
			powerS.write("ACILIM 5.1") #Set output current limit to 8A
			powerS.write("OUT ON") #Set output ON
			outPowerSCurr = powerS.write("IOUT?") #Query about output current

			time.sleep(5)

			# Record Output Voltage
			multimeter.write(":SENSe:FUNCtion 'VOLTage:DC'")  # Set the function to DC voltage
			multimeter.write(":SENSe:VOLTage:DC:RANGe:AUTO ON")  # Set the range to 10V, adjust as necessary
			x1y2 = multimeter.query(":READ?")
			x1y2 = x1y2[:13] + x1y2[14:]
			x1y2 = float(x1y2)
			#print(x1y2 + "V")
			c1 = sheet.cell(row = 28, column = 8)
			c1.value = x1y2

			loadS.write("CURRent 10")

			time.sleep(5)

			# Record Output Voltage
			multimeter.write(":SENSe:FUNCtion 'VOLTage:DC'")  # Set the function to DC voltage
			multimeter.write(":SENSe:VOLTage:DC:RANGe:AUTO ON")  # Set the range to 10V, adjust as necessary
			x1y2 = multimeter.query(":READ?")
			x1y2 = x1y2[:13] + x1y2[14:]
			x1y2 = float(x1y2)
			#print(x1y2 + "V")
			c1 = sheet.cell(row = 28, column = 11)
			c1.value = x1y2			

			# Input = 200V
			powerS.write("ACDC AC") #Set output to DC
			powerS.write("RANGE 200") #Set range to 200V mode
			powerS.write("ACVSET 200") #Set output voltage to 200V
			powerS.write("ACILIM 5.1") #Set output current limit to 8A
			powerS.write("OUT ON") #Set output ON
			outPowerSCurr = powerS.write("IOUT?") #Query about output current

			time.sleep(5)

			# Record Output Voltage
			multimeter.write(":SENSe:FUNCtion 'VOLTage:DC'")  # Set the function to DC voltage
			multimeter.write(":SENSe:VOLTage:DC:RANGe:AUTO ON")  # Set the range to 10V, adjust as necessary
			x1y2 = multimeter.query(":READ?")
			x1y2 = x1y2[:13] + x1y2[14:]
			x1y2 = float(x1y2)
			#print(x1y2 + "V")
			c1 = sheet.cell(row = 27, column = 11)
			c1.value = x1y2			

			# Input = 180V
			powerS.write("ACDC AC") #Set output to DC
			powerS.write("RANGE 200") #Set range to 200V mode
			powerS.write("ACVSET 180") #Set output voltage to 200V
			powerS.write("ACILIM 5.1") #Set output current limit to 8A
			powerS.write("OUT ON") #Set output ON
			outPowerSCurr = powerS.write("IOUT?") #Query about output current

			time.sleep(5)

			# Record Output Voltage
			multimeter.write(":SENSe:FUNCtion 'VOLTage:DC'")  # Set the function to DC voltage
			multimeter.write(":SENSe:VOLTage:DC:RANGe:AUTO ON")  # Set the range to 10V, adjust as necessary
			x1y2 = multimeter.query(":READ?")
			x1y2 = x1y2[:13] + x1y2[14:]
			x1y2 = float(x1y2)
			#print(x1y2 + "V")
			c1 = sheet.cell(row = 26, column = 11)
			c1.value = x1y2			

			# Input = 200V
			powerS.write("ACDC AC") #Set output to DC
			powerS.write("RANGE 200") #Set range to 200V mode
			powerS.write("ACVSET 200") #Set output voltage to 200V
			powerS.write("ACILIM 5.1") #Set output current limit to 8A
			powerS.write("OUT ON") #Set output ON
			outPowerSCurr = powerS.write("IOUT?") #Query about output current

		elif self.command_index == 17:
			print("Running internal code for step 17: nibel")
		elif self.command_index == 18:

			print("Running internal code for step 18: Does something")

			# Load Current = 12A, Find Voltage
			loadS.write("INPut ON")
			loadS.write("OUTput ON")
			loadS.write("CURRent 12")
			vol = loadS.query("MEASure:VOLTage?")
			vol = vol[:13] + vol[14:]
			vol = float(vol)

			# Set Load Conductance
			cond = 12.0/vol
			loadS.write("COND {:.8f}".format(cond))
			time.sleep(3)

			# CC -> CR
			loadS.write("INPut OFF")
			loadS.write("OUTput OFF")
			time.sleep(3)
			loadS.write("FUNC CR")
			time.sleep(3)
			loadS.write("INPut ON")
			loadS.write("OUTput ON")

			# Record Output Voltage
			multimeter.write(":SENSe:FUNCtion 'VOLTage:DC'")  # Set the function to DC voltage
			multimeter.write(":SENSe:VOLTage:DC:RANGe:AUTO ON")  # Set the range to 10V, adjust as necessary
			liveVolt = multimeter.query(":READ?")
			liveVolt = liveVolt[:13] + liveVolt[14:]
			liveVolt = float(liveVolt)
			c1 = sheet.cell(row = 27, column = 22)
			c1.value = liveVolt	

			## Repeat until output voltage is within 0.05 volts of 24V
			while not(liveVolt >23.95 and liveVolt < 24.05):
				# Measure Output Voltage
				multimeter.write(":SENSe:FUNCtion 'VOLTage:DC'")  # Set the function to DC voltage
				multimeter.write(":SENSe:VOLTage:DC:RANGe:AUTO ON")  # Set the range to 10V, adjust as necessary
				liveVolt = multimeter.query(":READ?")
				time.sleep(1)
				liveVolt = liveVolt[:13] + liveVolt[14:]
				liveVolt = float(liveVolt)

			# Record Load Current
			liveCur = loadS.query("MEASure:CURRent?")
			liveCur = float(liveCur)
			c1 = sheet.cell(row = 27, column = 22)
			c1.value = liveCur	

		elif self.command_index == 19:

			print("Running internal code for step 19")
			
			# CR -> CC
			time.sleep(3)
			loadS.write("INPut OFF")
			loadS.write("OUTput OFF")
			time.sleep(3)
			loadS.write("FUNC CC")
			loadS.write("CURRent 10")
			time.sleep(3)
			loadS.write("INPut ON")
			loadS.write("OUTput ON")

			# Record Output Voltage
			multimeter.write(":SENSe:FUNCtion 'VOLTage:DC'")  # Set the function to DC voltage
			multimeter.write(":SENSe:VOLTage:DC:RANGe:AUTO ON")  # Set the range to 10V, adjust as necessary
			x1y2 = multimeter.query(":READ?")
			x1y2 = x1y2[:13] + x1y2[14:]
			x1y2 = float(x1y2)
			#print(x1y2 + "V")
			c1 = sheet.cell(row = 35, column = 14)
			c1.value = x1y2

			# Load Off
			loadS.write("INPut OFF")
			loadS.write("OUTput OFF")
			
		elif self.command_index == 20:

			print("Running internal code for step 20")
			print("REAL Running internal code for step 20")
			
			# Record Output Voltage
			multimeter.write(":SENSe:FUNCtion 'VOLTage:DC'")  # Set the function to DC voltage
			multimeter.write(":SENSe:VOLTage:DC:RANGe:AUTO ON")  # Set the range to 10V, adjust as necessary
			x1y2 = multimeter.query(":READ?")
			x1y2 = x1y2[:13] + x1y2[14:]
			x1y2 = float(x1y2)
			#print(x1y2 + "V")
			c1 = sheet.cell(row = 34, column = 11)
			c1.value = x1y2

			# Reset Power, Load Off
			#powerS.write("OUTput OFF")
			powerS.write("*RST")
			loadS.write("INPut OFF")
			loadS.write("OUTput OFF")

		##### Internal Code End #####
		
		self.loading_window.destroy()
		self.label.config(text=self.commands[self.language][self.command_index])
		self.entry.pack()

	def finish(self):
		self.label.config(text="Commands Completed" if self.language == 'English' else "コマンド完了", font=("Arial", 20))
		self.entry.pack_forget()
		print("Collected Data:", self.data)

		# Save the workbook
		wb.save(r"BatteryCharger.xlsx")



		# Close the application
		self.master.quit()

if __name__ == "__main__":
	root = tk.Tk()
	app = CommandGUI(root)
	root.mainloop()
	 # Extract the directory and filename
	dir_name = os.path.dirname("BatteryCharger.xlsx")
	base_name = os.path.basename("BatteryCharger.xlsx")

	# Create the new file name by appending '_copy' before the extension
	name, ext = os.path.splitext(base_name)
	serial = sheet.cell(row=9, column=5).value
	new_file_path = os.path.join(dir_name, f"{name}_{serial}{ext}")

	# Copy the file
	shutil.copy2("BatteryCharger.xlsx", new_file_path)
	print(f"File copied to {new_file_path}")